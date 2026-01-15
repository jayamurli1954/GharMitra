"""
Export Utilities for Excel and PDF Generation
Provides functions to export reports to Excel and PDF formats
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Any, Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT


from decimal import Decimal

def create_maintenance_bill_pdf(
    bill_data: Dict[str, Any],
    society_info: Dict[str, Any]
) -> BytesIO:
    """
    Create a professional maintenance bill PDF
    
    Args:
        bill_data: Bill information (flat, amount, breakdown, etc.)
        society_info: Society information (name, address, logo_url)
        
    Returns:
        BytesIO: PDF file in memory
    """
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    # Watermark for draft bills
    is_draft = not bill_data.get('is_posted', True) # Default to True (posted) if not provided
    
    def on_page(canvas, doc):
        if is_draft:
            canvas.saveState()
            canvas.setFont('Helvetica-Bold', 100)
            canvas.setFillGray(0.5, 0.2) # Gray color with 20% alpha
            canvas.translate(A4[0]/2, A4[1]/2)
            canvas.rotate(45)
            canvas.drawCentredString(0, 0, "DRAFT")
            canvas.restoreState()

    # Styles
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.grey,
        spaceAfter=12,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=13,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=8,
        spaceBefore=12,
        fontName='Helvetica-Bold'
    )
    
    # Helper to format amounts safely
    def format_val(val):
        if val is None: return "0.00"
        try:
            return f"{float(val):.2f}"
        except (ValueError, TypeError):
            return str(val)

    # Build content
    content = []
    
    # Society Header
    content.append(Paragraph(society_info.get('name', 'Housing Society'), title_style))
    if society_info.get('address'):
        content.append(Paragraph(society_info['address'], subtitle_style))
    content.append(Spacer(1, 0.2 * inch))
    
    # Bill Title
    try:
        month_name = datetime.strptime(f"{bill_data.get('year')}-{bill_data.get('month'):02d}-01", "%Y-%m-%d").strftime("%B %Y")
    except:
        month_name = f"Month: {bill_data.get('month')} Year: {bill_data.get('year')}"
    
    bill_title = f"<b>MAINTENANCE BILL</b><br/>{month_name}"
    content.append(Paragraph(bill_title, heading_style))
    content.append(Spacer(1, 0.15 * inch))
    
    # Bill Information
    bill_info_data = [
        ['Bill Number:', bill_data.get('bill_number', 'N/A')],
        ['Flat Number:', bill_data.get('flat_number', 'N/A')],
        ['Member Name:', bill_data.get('member_name', 'N/A')],
        ['Generated Date:', datetime.fromisoformat(bill_data.get('created_at', datetime.now().isoformat())).strftime('%d %b %Y')],
        ['Status:', 'DRAFT' if is_draft else bill_data.get('status', 'Unpaid').upper()],
    ]
    
    bill_info_table = Table(bill_info_data, colWidths=[2*inch, 3.5*inch])
    bill_info_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#444444')),
        ('TEXTCOLOR', (1, 0), (1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LINEBELOW', (0, -1), (-1, -1), 0.5, colors.grey),
    ]))
    content.append(bill_info_table)
    content.append(Spacer(1, 0.25 * inch))
    
    # Breakdown Heading
    content.append(Paragraph('Bill Breakdown', heading_style))
    
    # Breakdown table
    breakdown = bill_data.get('breakdown', {})
    breakdown_data = [['Description', 'Amount (₹)']]
    
    # Fixed expenses
    if 'fixed_expenses' in breakdown and breakdown['fixed_expenses']:
        fixed_list = breakdown.get('fixed_expenses_list', [])
        if fixed_list:
            for expense in fixed_list:
                breakdown_data.append([
                    f"  • {expense.get('name', 'Fixed Expense')}",
                    format_val(expense.get('amount'))
                ])
        else:
            breakdown_data.append(['Fixed Expenses', format_val(breakdown['fixed_expenses'])])
    
    # Water charges
    if 'water_charges' in breakdown and breakdown['water_charges']:
        water_exp = breakdown.get('water_expenses', {})
        if water_exp and 'flat_occupants' in water_exp:
            breakdown_data.append([
                f"Water Charges ({water_exp.get('flat_occupants', 0)} occupants × ₹{format_val(water_exp.get('per_person_charge', 0))})",
                format_val(breakdown['water_charges'])
            ])
        else:
            breakdown_data.append(['Water Charges', format_val(breakdown['water_charges'])])
    
    # Sinking fund
    if 'sinking_fund' in breakdown and breakdown['sinking_fund']:
        breakdown_data.append(['Sinking Fund', format_val(breakdown['sinking_fund'])])
    
    # Sqft calculation (for sqft-based billing)
    if 'sqft_calculation' in breakdown and breakdown['sqft_calculation']:
        # If it's a string like "1000 sq ft x ₹2 = ₹2000", just add it
        breakdown_data.append(['Square Feet Calculation', str(breakdown['sqft_calculation'])])
    
    # Total amount row
    breakdown_data.append(['', ''])  # Spacer
    breakdown_data.append(['TOTAL AMOUNT', format_val(bill_data.get('amount', 0))])
    
    breakdown_table = Table(breakdown_data, colWidths=[4*inch, 1.5*inch])
    breakdown_table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 11),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        
        # Data rows
        ('FONT', (0, 1), (-1, -3), 'Helvetica', 10),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Gridlines
        ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
        ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#1f4788')),
        
        # Total row (bold and highlighted)
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 12),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1f4788')),
        
        # Padding
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
    ]))
    content.append(breakdown_table)
    content.append(Spacer(1, 0.3 * inch))
    
    # Payment status
    if not is_draft and bill_data.get('status') == 'paid' and bill_data.get('paid_at'):
        try:
            paid_date = datetime.fromisoformat(bill_data['paid_at']).strftime('%d %b %Y')
        except:
            paid_date = str(bill_data['paid_at'])
        payment_method = bill_data.get('payment_method', 'Not specified')
        payment_info = f"<b>Payment Status:</b> PAID on {paid_date}<br/><b>Payment Method:</b> {payment_method}"
        payment_style = ParagraphStyle(
            'PaymentInfo',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.HexColor('#28a745'),
            spaceAfter=12,
            leftIndent=10
        )
        content.append(Paragraph(payment_info, payment_style))
    elif not is_draft:
        # Payment instructions for unpaid bills
        payment_instructions = """
        <b>Payment Instructions:</b><br/>
        Please make the payment by bank transfer or cash to the society office.<br/>
        For bank transfer, use the society's registered bank account.<br/>
        Keep the receipt for your records.
        """
        payment_style = ParagraphStyle(
            'PaymentInstructions',
            parent=styles['Normal'],
            fontSize=9,
            textColor=colors.grey,
            spaceAfter=12,
            leftIndent=10
        )
        content.append(Paragraph(payment_instructions, payment_style))
    else:
        # Draft message
        draft_msg = "<b>This is a DRAFT bill.</b><br/>Please wait for final posting before making payment."
        draft_style = ParagraphStyle(
            'DraftInfo',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.red,
            spaceAfter=12,
            leftIndent=10
        )
        content.append(Paragraph(draft_msg, draft_style))
    
    content.append(Spacer(1, 0.3 * inch))
    
    # Footer
    footer_text = f"""
    <i>This is a computer-generated bill. Generated on {datetime.now().strftime('%d %b %Y at %I:%M %p')}</i><br/>
    <i>For any queries, please contact the society office.</i>
    """
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    content.append(Paragraph(footer_text, footer_style))
    
    # Build PDF
    doc.build(content, onFirstPage=on_page, onLaterPages=on_page)
    buffer.seek(0)
    return buffer


class ExcelExporter:
    """Excel export functionality using openpyxl"""
    
    @staticmethod
    def create_general_ledger_excel(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """
        Create Excel file for General Ledger Report
        
        Args:
            report_data: Report data dictionary
            society_info: Society information (name, address, logo_url)
            
        Returns:
            BytesIO: Excel file in memory
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "General Ledger"
        
        # Styles
        header_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=12, bold=True)
        subheader_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = society_info.get('name', 'Society Name')
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        cell.value = "GENERAL LEDGER REPORT"
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        cell = ws[f'A{row}']
        period = f"Period: {report_data.get('from_date', '')} to {report_data.get('to_date', '')}"
        cell.value = period
        cell.font = normal_font
        cell.alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Column headers
        headers = ['Account', 'Date', 'Description', 'Debit', 'Credit', 'Balance']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row += 1
        
        # Data
        ledger_data = report_data.get('ledger', {})
        
        for account_code, account_info in ledger_data.items():
            # Account header
            ws.merge_cells(f'A{row}:F{row}')
            cell = ws[f'A{row}']
            cell.value = f"{account_code} - {account_info.get('account_name', '')}"
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            cell.border = border
            row += 1
            
            # Opening balance
            cell = ws.cell(row=row, column=1)
            cell.value = "Opening Balance"
            cell.font = normal_font
            cell.border = border
            
            opening_balance = account_info.get('opening_balance', 0)
            cell = ws.cell(row=row, column=6)
            cell.value = opening_balance
            cell.number_format = '#,##0.00'
            cell.font = normal_font
            cell.border = border
            row += 1
            
            # Transactions
            for txn in account_info.get('transactions', []):
                cell = ws.cell(row=row, column=2)
                cell.value = txn.get('date', '')
                cell.font = normal_font
                cell.border = border
                
                cell = ws.cell(row=row, column=3)
                cell.value = txn.get('description', '')
                cell.font = normal_font
                cell.border = border
                
                cell = ws.cell(row=row, column=4)
                cell.value = txn.get('debit', 0)
                cell.number_format = '#,##0.00'
                cell.font = normal_font
                cell.border = border
                
                cell = ws.cell(row=row, column=5)
                cell.value = txn.get('credit', 0)
                cell.number_format = '#,##0.00'
                cell.font = normal_font
                cell.border = border
                
                cell = ws.cell(row=row, column=6)
                cell.value = txn.get('balance', 0)
                cell.number_format = '#,##0.00'
                cell.font = normal_font
                cell.border = border
                
                row += 1
            
            # Totals
            cell = ws.cell(row=row, column=1)
            cell.value = "TOTAL"
            cell.font = subheader_font
            cell.fill = total_fill
            cell.border = border
            
            cell = ws.cell(row=row, column=4)
            cell.value = account_info.get('total_debit', 0)
            cell.number_format = '#,##0.00'
            cell.font = subheader_font
            cell.fill = total_fill
            cell.border = border
            
            cell = ws.cell(row=row, column=5)
            cell.value = account_info.get('total_credit', 0)
            cell.number_format = '#,##0.00'
            cell.font = subheader_font
            cell.fill = total_fill
            cell.border = border
            
            # Closing balance
            cell = ws.cell(row=row, column=6)
            cell.value = account_info.get('closing_balance', 0)
            cell.number_format = '#,##0.00'
            cell.font = subheader_font
            cell.fill = total_fill
            cell.border = border
            
            row += 2
        
        # Auto-adjust column widths
        for col in range(1, 7):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 18
        
        ws.column_dimensions['C'].width = 40  # Description column
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    @staticmethod
    def create_trial_balance_excel(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create Excel file for Trial Balance"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trial Balance"
        
        # Styles
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = society_info.get('name', 'Society Name')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"TRIAL BALANCE AS ON {report_data.get('as_on_date', '')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Column headers
        headers = ['Account Code', 'Account Name', 'Debit Balance', 'Credit Balance']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
        # Data
        row = 5
        for item in report_data.get('items', []):
            ws.cell(row=row, column=1, value=item.get('account_code', '')).border = border
            ws.cell(row=row, column=2, value=item.get('account_name', '')).border = border
            
            d_cell = ws.cell(row=row, column=3, value=item.get('debit_balance', 0))
            d_cell.number_format = '#,##0.00'
            d_cell.border = border
            
            c_cell = ws.cell(row=row, column=4, value=item.get('credit_balance', 0))
            c_cell.number_format = '#,##0.00'
            c_cell.border = border
            row += 1
            
        # Total row
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2).fill = total_fill
        ws.cell(row=row, column=2).border = border
        
        td_cell = ws.cell(row=row, column=3, value=report_data.get('total_debit', 0))
        td_cell.font = Font(bold=True)
        td_cell.number_format = '#,##0.00'
        td_cell.fill = total_fill
        td_cell.border = border
        
        tc_cell = ws.cell(row=row, column=4, value=report_data.get('total_credit', 0))
        tc_cell.font = Font(bold=True)
        tc_cell.number_format = '#,##0.00'
        tc_cell.fill = total_fill
        tc_cell.border = border
        
        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 18
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def create_income_and_expenditure_excel(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create Excel file for Income & Expenditure"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Income & Expenditure"
        
        # Styles
        header_font = Font(name='Arial', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Header
        ws.merge_cells('A1:C1')
        ws['A1'] = society_info.get('name', 'Society Name')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:C2')
        ws['A2'] = f"INCOME & EXPENDITURE FOR {report_data.get('from_date', '')} TO {report_data.get('to_date', '')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        # Summary
        row = 4
        ws.cell(row=row, column=1, value="Total Income:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_data.get('total_income', 0)).number_format = '#,##0.00'
        row += 1
        ws.cell(row=row, column=1, value="Total Expenditure:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=report_data.get('total_expenditure', 0)).number_format = '#,##0.00'
        row += 1
        net_income = report_data.get('net_income', 0)
        cell = ws.cell(row=row, column=1, value="Net Surplus/(Deficit):")
        cell.font = Font(bold=True)
        val_cell = ws.cell(row=row, column=2, value=net_income)
        val_cell.font = Font(bold=True)
        val_cell.number_format = '#,##0.00'
        
        row += 2
        
        # Details
        def add_section(title, items):
            nonlocal row
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
            row += 1
            headers = ['Code', 'Account Name', 'Amount']
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row, column=col)
                c.value = h
                c.font = header_font
                c.fill = header_fill
                c.border = border
            row += 1
            for item in items:
                ws.cell(row=row, column=1, value=item.get('account_code', '')).border = border
                ws.cell(row=row, column=2, value=item.get('account_name', '')).border = border
                amt_cell = ws.cell(row=row, column=3, value=item.get('amount', 0))
                amt_cell.number_format = '#,##0.00'
                amt_cell.border = border
                row += 1
            row += 1

        add_section("INCOME", report_data.get('income_items', []))
        add_section("EXPENDITURE", report_data.get('expenditure_items', []))
        
        # Adjust widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def create_balance_sheet_excel(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create Excel file for Balance Sheet"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balance Sheet"
        
        # Header
        ws.merge_cells('A1:D1')
        ws['A1'] = society_info.get('name', 'Society Name')
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        ws.merge_cells('A2:D2')
        ws['A2'] = f"BALANCE SHEET AS ON {report_data.get('as_on_date', '')}"
        ws['A2'].alignment = Alignment(horizontal='center')
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        row = 4
        # Liabilities & Assets side by side-ish or sequential
        # Let's do sequential for simplicity in Excel
        
        def add_bs_section(title, items, total_label, total_val):
            nonlocal row
            ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=12)
            row += 1
            ws.cell(row=row, column=1, value="Account").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=1).border = border
            ws.cell(row=row, column=2, value="Amount").font = header_font
            ws.cell(row=row, column=2).fill = header_fill
            ws.cell(row=row, column=2).border = border
            row += 1
            for item in items:
                ws.cell(row=row, column=1, value=item.get('account_name', '')).border = border
                a_cell = ws.cell(row=row, column=2, value=item.get('balance', 0))
                a_cell.number_format = '#,##0.00'
                a_cell.border = border
                row += 1
            ws.cell(row=row, column=1, value=total_label).font = Font(bold=True)
            ws.cell(row=row, column=1).border = border
            t_cell = ws.cell(row=row, column=2, value=total_val)
            t_cell.font = Font(bold=True)
            t_cell.number_format = '#,##0.00'
            t_cell.border = border
            row += 2

        add_bs_section("ASSETS", report_data.get('assets', []), "Total Assets", report_data.get('total_assets', 0))
        # Combine Liabilities and Capital
        liab_cap = report_data.get('liabilities', []) + report_data.get('capital', [])
        add_bs_section("LIABILITIES & CAPITAL", liab_cap, "Total Liabilities & Capital", report_data.get('total_liabilities_capital', 0))

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def create_simple_report_excel(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any],
        report_title: str,
        columns: List[str],
        data_key: str
    ) -> BytesIO:
        """
        Create Excel file for simple table reports
        
        Args:
            report_data: Report data dictionary
            society_info: Society information
            report_title: Title of the report
            columns: Column headers
            data_key: Key in report_data containing the table data
            
        Returns:
            BytesIO: Excel file in memory
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = report_title[:31]  # Excel sheet name limit
        
        # Styles
        header_font = Font(name='Arial', size=14, bold=True)
        title_font = Font(name='Arial', size=12, bold=True)
        subheader_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Header
        row = 1
        ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
        cell = ws[f'A{row}']
        cell.value = society_info.get('name', 'Society Name')
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        
        row += 1
        ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
        cell = ws[f'A{row}']
        cell.value = report_title.upper()
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center')
        
        if report_data.get('from_date') and report_data.get('to_date'):
            row += 1
            ws.merge_cells(f'A{row}:{get_column_letter(len(columns))}{row}')
            cell = ws[f'A{row}']
            period = f"Period: {report_data.get('from_date')} to {report_data.get('to_date')}"
            cell.value = period
            cell.font = normal_font
            cell.alignment = Alignment(horizontal='center')
        
        row += 2
        
        # Column headers
        for col_num, header in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col_num)
            cell.value = header
            cell.font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = border
        
        row += 1
        
        # Data
        data = report_data.get(data_key, [])
        for item in data:
            for col_num, col_name in enumerate(columns, 1):
                cell = ws.cell(row=row, column=col_num)
                # Try to get value by lowercase key with underscores
                key = col_name.lower().replace(' ', '_')
                value = item.get(key, item.get(col_name, ''))
                cell.value = value
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.font = normal_font
                cell.border = border
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(columns) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 18
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class PDFExporter:
    """PDF export functionality using ReportLab"""
    
    @staticmethod
    def create_payment_receipt_pdf(
        receipt_number: str,
        payment_date: Any,
        payment_mode: str,
        amount: float,
        late_fee: float,
        transaction_reference: Optional[str],
        bill_number: str,
        bill_date: Any,
        billing_period: str,
        member_name: str,
        flat_number: str,
        society_name: str,
        society_address: Optional[str],
        maintenance_amount: float,
        sinking_fund: float,
        other_charges: float
    ) -> BytesIO:
        """
        Create PDF payment receipt
        
        Args:
            receipt_number: Receipt number
            payment_date: Date of payment
            payment_mode: Mode of payment (cash, cheque, UPI, etc.)
            amount: Payment amount
            late_fee: Late fee charged
            transaction_reference: Cheque no, UPI ref, etc.
            bill_number: Original bill number
            bill_date: Bill date
            billing_period: Billing period (MM/YYYY)
            member_name: Member name
            flat_number: Flat number
            society_name: Society name
            society_address: Society address
            maintenance_amount: Maintenance amount
            sinking_fund: Sinking fund amount
            other_charges: Other charges
            
        Returns:
            BytesIO: PDF file in memory
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'ReceiptTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#2E7D32'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'ReceiptHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#1976D2'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        )
        
        # Header: Society Name
        society_para = Paragraph(society_name, title_style)
        elements.append(society_para)
        
        if society_address:
            address_para = Paragraph(society_address, ParagraphStyle(
                'Address',
                parent=styles['Normal'],
                fontSize=10,
                alignment=TA_CENTER,
                spaceAfter=20
            ))
            elements.append(address_para)
        
        # Receipt Title
        receipt_title = Paragraph("PAYMENT RECEIPT", heading_style)
        receipt_title.alignment = TA_CENTER
        elements.append(receipt_title)
        elements.append(Spacer(1, 0.2*inch))
        
        # Receipt Info Table
        receipt_info = [
            ['Receipt No:', receipt_number, 'Date:', str(payment_date)],
            ['Bill No:', bill_number, 'Bill Date:', str(bill_date)],
            ['Member Name:', member_name, 'Flat No:', flat_number],
            ['Billing Period:', billing_period, '', '']
        ]
        
        receipt_table = Table(receipt_info, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
        receipt_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
            ('TEXTCOLOR', (2, 0), (2, -1), colors.HexColor('#1976D2')),
        ]))
        elements.append(receipt_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Payment Details Heading
        payment_heading = Paragraph("PAYMENT DETAILS", ParagraphStyle(
            'PaymentHeading',
            parent=styles['Heading3'],
            fontSize=12,
            textColor=colors.HexColor('#1976D2'),
            spaceAfter=10,
            fontName='Helvetica-Bold'
        ))
        elements.append(payment_heading)
        
        # Payment Details Table
        total_bill = maintenance_amount + sinking_fund + other_charges
        total_with_late_fee = amount + late_fee
        
        payment_data = [
            ['Description', 'Amount'],
            ['Maintenance Charges', f'₹{maintenance_amount:,.2f}'],
            ['Sinking Fund', f'₹{sinking_fund:,.2f}'],
            ['Other Charges', f'₹{other_charges:,.2f}'],
            ['', ''],
            ['Bill Amount', f'₹{total_bill:,.2f}'],
        ]
        
        if late_fee > 0:
            payment_data.append(['Late Fee', f'₹{late_fee:,.2f}'])
            payment_data.append(['Total Amount Paid', f'₹{total_with_late_fee:,.2f}'])
        else:
            payment_data.append(['Total Amount Paid', f'₹{amount:,.2f}'])
        
        payment_table = Table(payment_data, colWidths=[4*inch, 2.5*inch])
        payment_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -3), 0.5, colors.grey),
            ('LINEABOVE', (0, -2), (-1, -2), 2, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E3F2FD')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#C8E6C9')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#2E7D32')),
        ]))
        elements.append(payment_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Payment Mode Info
        mode_info = [
            ['Payment Mode:', payment_mode.upper()]
        ]
        if transaction_reference:
            mode_info.append(['Transaction Reference:', transaction_reference])
        
        mode_table = Table(mode_info, colWidths=[2*inch, 4.5*inch])
        mode_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor('#1976D2')),
        ]))
        elements.append(mode_table)
        elements.append(Spacer(1, 0.5*inch))
        
        # Footer Note
        note = Paragraph(
            "<i>This is a computer-generated receipt and does not require a signature.</i>",
            ParagraphStyle(
                'Note',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=TA_CENTER
            )
        )
        elements.append(note)
        
        # Thank you message
        thank_you = Paragraph(
            "<b>Thank you for your payment!</b>",
            ParagraphStyle(
                'ThankYou',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.HexColor('#2E7D32'),
                alignment=TA_CENTER,
                spaceAfter=20
            )
        )
        elements.append(Spacer(1, 0.2*inch))
        elements.append(thank_you)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer
    
    @staticmethod
    def create_general_ledger_pdf(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """
        Create PDF file for General Ledger Report
        
        Args:
            report_data: Report data dictionary
            society_info: Society information (name, address, logo_url)
            
        Returns:
            BytesIO: PDF file in memory
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        # Container for the 'Flowable' objects
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=12,
            textColor=colors.HexColor('#366092'),
            spaceAfter=6,
            alignment=TA_CENTER
        )
        
        # Society name
        society_name = Paragraph(society_info.get('name', 'Society Name'), title_style)
        elements.append(society_name)
        
        # Report title
        title = Paragraph("GENERAL LEDGER REPORT", heading_style)
        elements.append(title)
        
        # Period
        period = f"Period: {report_data.get('from_date', '')} to {report_data.get('to_date', '')}"
        period_para = Paragraph(period, styles['Normal'])
        period_para.alignment = TA_CENTER
        elements.append(period_para)
        elements.append(Spacer(1, 0.2*inch))
        
        # Ledger data
        ledger_data = report_data.get('ledger', {})
        
        for account_code, account_info in ledger_data.items():
            # Account header
            account_title = f"{account_code} - {account_info.get('account_name', '')}"
            account_para = Paragraph(f"<b>{account_title}</b>", styles['Heading3'])
            elements.append(account_para)
            elements.append(Spacer(1, 0.1*inch))
            
            # Table data
            table_data = [
                ['Date', 'Description', 'Debit', 'Credit', 'Balance']
            ]
            
            # Opening balance
            opening_balance = account_info.get('opening_balance', 0)
            table_data.append(['', 'Opening Balance', '', '', f"₹{opening_balance:,.2f}"])
            
            # Transactions
            for txn in account_info.get('transactions', []):
                table_data.append([
                    txn.get('date', ''),
                    txn.get('description', ''),
                    f"₹{txn.get('debit', 0):,.2f}" if txn.get('debit', 0) > 0 else '',
                    f"₹{txn.get('credit', 0):,.2f}" if txn.get('credit', 0) > 0 else '',
                    f"₹{txn.get('balance', 0):,.2f}"
                ])
            
            # Totals
            table_data.append([
                '',
                'TOTAL',
                f"₹{account_info.get('total_debit', 0):,.2f}",
                f"₹{account_info.get('total_credit', 0):,.2f}",
                f"₹{account_info.get('closing_balance', 0):,.2f}"
            ])
            
            # Create table
            table = Table(table_data, colWidths=[1*inch, 3*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#D9E1F2')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def create_trial_balance_pdf(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create a professional Trial Balance PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, margin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        elements.append(Paragraph(society_info.get('name', 'Society Name'), styles['Title']))
        elements.append(Paragraph(f"TRIAL BALANCE AS ON {report_data.get('as_on_date', '')}", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Table data
        table_data = [['Account Code', 'Account Name', 'Debit Balance', 'Credit Balance']]
        for item in report_data.get('items', []):
            table_data.append([
                item.get('account_code', ''),
                item.get('account_name', ''),
                f"₹{item.get('debit_balance', 0):,.2f}" if item.get('debit_balance', 0) > 0 else '-',
                f"₹{item.get('credit_balance', 0):,.2f}" if item.get('credit_balance', 0) > 0 else '-'
            ])
            
        # Total row
        table_data.append([
            '', 'TOTAL',
            f"₹{report_data.get('total_debit', 0):,.2f}",
            f"₹{report_data.get('total_credit', 0):,.2f}"
        ])
        
        table = Table(table_data, colWidths=[1*inch, 3*inch, 1.25*inch, 1.25*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#f0f0f0')),
        ]))
        elements.append(table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_income_and_expenditure_pdf(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create a professional Income & Expenditure PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, margin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        elements.append(Paragraph(society_info.get('name', 'Society Name'), styles['Title']))
        elements.append(Paragraph(f"INCOME & EXPENDITURE ACCOUNT FOR {report_data.get('from_date', '')} TO {report_data.get('to_date', '')}", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary Row (Income vs Expenditure)
        summary_data = [
            [f"Total Income: ₹{report_data.get('total_income', 0):,.2f}", f"Total Expenditure: ₹{report_data.get('total_expenditure', 0):,.2f}"]
        ]
        summary_table = Table(summary_data, colWidths=[3.25*inch, 3.25*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 12),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'RIGHT'),
            ('TEXTCOLOR', (0,0), (0,0), colors.HexColor('#2e7d32')), # Green for income
            ('TEXTCOLOR', (1,0), (1,0), colors.HexColor('#c62828')), # Red for expenditure
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.1*inch))
        
        # Net Surplus/Deficit
        net_income = report_data.get('net_income', 0)
        net_text = f"NET SURPLUS: ₹{net_income:,.2f}" if net_income >= 0 else f"NET DEFICIT: ₹{abs(net_income):,.2f}"
        net_color = colors.HexColor('#2e7d32') if net_income >= 0 else colors.HexColor('#c62828')
        elements.append(Paragraph(f"<b>{net_text}</b>", ParagraphStyle('NetStyle', parent=styles['Normal'], fontSize=14, textColor=net_color, alignment=TA_CENTER)))
        elements.append(Spacer(1, 0.3*inch))
        
        # Details (Income)
        elements.append(Paragraph("<b>INCOME DETAILS</b>", styles['Heading3']))
        income_data = [['Account Code', 'Account Name', 'Amount (₹)']]
        for item in report_data.get('income_items', []):
            income_data.append([item.get('account_code', ''), item.get('account_name', ''), f"{item.get('amount', 0):,.2f}"])
        
        if len(income_data) > 1:
            i_table = Table(income_data, colWidths=[1*inch, 4*inch, 1.5*inch])
            i_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ]))
            elements.append(i_table)
        else:
            elements.append(Paragraph("No income recorded for this period.", styles['Normal']))
            
        elements.append(Spacer(1, 0.3*inch))
        
        # Details (Expenditure)
        elements.append(Paragraph("<b>EXPENDITURE DETAILS</b>", styles['Heading3']))
        exp_data = [['Account Code', 'Account Name', 'Amount (₹)']]
        for item in report_data.get('expenditure_items', []):
            exp_data.append([item.get('account_code', ''), item.get('account_name', ''), f"{item.get('amount', 0):,.2f}"])
            
        if len(exp_data) > 1:
            e_table = Table(exp_data, colWidths=[1*inch, 4*inch, 1.5*inch])
            e_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ]))
            elements.append(e_table)
        else:
            elements.append(Paragraph("No expenditure recorded for this period.", styles['Normal']))
            
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_balance_sheet_pdf(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """Create a professional Balance Sheet PDF"""
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, margin=0.5*inch)
        elements = []
        styles = getSampleStyleSheet()
        
        # Header
        elements.append(Paragraph(society_info.get('name', 'Society Name'), styles['Title']))
        elements.append(Paragraph(f"BALANCE SHEET AS ON {report_data.get('as_on_date', '')}", styles['Heading2']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Assets Table
        elements.append(Paragraph("<b>ASSETS</b>", styles['Heading3']))
        asset_data = [['Account Name', 'Amount (₹)']]
        for item in report_data.get('assets', []):
            asset_data.append([item.get('account_name', ''), f"{item.get('balance', 0):,.2f}"])
        
        asset_table = Table(asset_data, colWidths=[5*inch, 1.5*inch])
        asset_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(asset_table)
        elements.append(Paragraph(f"<b>Total Assets: ₹{report_data.get('total_assets', 0):,.2f}</b>", ParagraphStyle('TotalStyle', parent=styles['Normal'], alignment=TA_RIGHT)))
        elements.append(Spacer(1, 0.3*inch))
        
        # Liabilities & Capital Table
        elements.append(Paragraph("<b>LIABILITIES & CAPITAL</b>", styles['Heading3']))
        liab_data = [['Account Name', 'Amount (₹)']]
        # Combine Liabilities and Capital
        all_liab = report_data.get('liabilities', []) + report_data.get('capital', [])
        for item in all_liab:
            liab_data.append([item.get('account_name', ''), f"{item.get('balance', 0):,.2f}"])
            
        liab_table = Table(liab_data, colWidths=[5*inch, 1.5*inch])
        liab_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(liab_table)
        elements.append(Paragraph(f"<b>Total Liabilities & Capital: ₹{report_data.get('total_liabilities_capital', 0):,.2f}</b>", ParagraphStyle('TotalStyle', parent=styles['Normal'], alignment=TA_RIGHT)))
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_simple_report_pdf(
        report_data: Dict[str, Any],
        society_info: Dict[str, Any],
        report_title: str,
        columns: List[str],
        data_key: str
    ) -> BytesIO:
        """
        Create PDF file for simple table reports
        
        Args:
            report_data: Report data dictionary
            society_info: Society information
            report_title: Title of the report
            columns: Column headers
            data_key: Key in report_data containing the table data
            
        Returns:
            BytesIO: PDF file in memory
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter,
                               rightMargin=0.5*inch, leftMargin=0.5*inch,
                               topMargin=0.5*inch, bottomMargin=0.5*inch)
        
        elements = []
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            alignment=TA_CENTER
        )
        
        # Society name
        society_name = Paragraph(society_info.get('name', 'Society Name'), title_style)
        elements.append(society_name)
        
        # Report title
        title = Paragraph(report_title.upper(), title_style)
        elements.append(title)
        
        # Period (if applicable)
        if report_data.get('from_date') and report_data.get('to_date'):
            period = f"Period: {report_data.get('from_date')} to {report_data.get('to_date')}"
            period_para = Paragraph(period, styles['Normal'])
            period_para.alignment = TA_CENTER
            elements.append(period_para)
        
        elements.append(Spacer(1, 0.2*inch))
        
        # Table data
        table_data = [columns]
        
        data = report_data.get(data_key, [])
        for item in data:
            row = []
            for col_name in columns:
                key = col_name.lower().replace(' ', '_')
                value = item.get(key, item.get(col_name, ''))
                if isinstance(value, (int, float)):
                    row.append(f"₹{value:,.2f}")
                else:
                    row.append(str(value))
            table_data.append(row)
        
        # Calculate column widths
        col_width = 7.0 / len(columns)
        col_widths = [col_width * inch] * len(columns)
        
        # Create table
        table = Table(table_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        return buffer

    @staticmethod
    def create_formal_receipt_pdf(
        voucher_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """
        Create a formal, member-friendly Receipt PDF
        Based on user redesign request
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=50,
            leftMargin=50,
            topMargin=50,
            bottomMargin=50
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Styles
        title_style = ParagraphStyle(
            'ReceiptTitle',
            parent=styles['Heading1'],
            fontSize=22, # Slightly smaller for better proportions
            textColor=colors.HexColor('#2e7d32'),
            alignment=TA_LEFT,
            fontName='Helvetica-Bold',
            leftIndent=0,
            spaceAfter=0
        )
        
        soc_name_style = ParagraphStyle(
            'SocietyName',
            parent=styles['Normal'],
            fontSize=14,
            fontName='Helvetica-Bold',
            alignment=TA_RIGHT
        )
        
        soc_details_style = ParagraphStyle(
            'SocietyDetails',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_RIGHT,
            textColor=colors.grey
        )
        
        label_style = ParagraphStyle(
            'LabelStyle',
            parent=styles['Normal'],
            fontSize=10,
            fontName='Helvetica-Bold'
        )
        
        val_style = ParagraphStyle(
            'ValueStyle',
            parent=styles['Normal'],
            fontSize=10
        )

        content_style = ParagraphStyle(
            'ContentStyle',
            parent=styles['Normal'],
            fontSize=11,
            leading=16,
            spaceBefore=5,
            spaceAfter=5
        )

        # Header Construction - Logo centered above society name
        logo_url = society_info.get('logo_url')
        logo = None
        if logo_url:
            try:
                import os
                import requests
                from reportlab.platypus import Image

                # Check if it's a local file path or URL
                if logo_url.startswith(('http://', 'https://')):
                    # Download from URL
                    response = requests.get(logo_url, timeout=5)
                    if response.status_code == 200:
                        logo = Image(BytesIO(response.content), width=0.8*inch, height=0.8*inch)
                elif os.path.exists(logo_url):
                    # Load from local file
                    logo = Image(logo_url, width=0.8*inch, height=0.8*inch)
                else:
                    logger.warning(f"Logo file not found: {logo_url}")
            except Exception as e:
                logger.warning(f"Could not load logo: {e}")

        # Society info with logo centered above
        if logo:
            # Center the logo
            logo_center_style = ParagraphStyle(
                'LogoCenter',
                parent=styles['Normal'],
                alignment=TA_CENTER
            )
            elements.append(logo)
            elements.append(Spacer(1, 0.1*inch))

        # Centered society name and address
        center_name_style = ParagraphStyle(
            'CenterName',
            parent=styles['Normal'],
            fontSize=14,
            fontName='Helvetica-Bold',
            alignment=TA_CENTER,
            spaceAfter=4
        )

        center_details_style = ParagraphStyle(
            'CenterDetails',
            parent=styles['Normal'],
            fontSize=9,
            alignment=TA_CENTER,
            textColor=colors.grey,
            spaceAfter=2
        )

        elements.append(Paragraph(society_info.get('name', '').upper(), center_name_style))
        if society_info.get('address'):
            elements.append(Paragraph(society_info.get('address', ''), center_details_style))
        if society_info.get('email'):
            elements.append(Paragraph(society_info.get('email', ''), center_details_style))
        if society_info.get('pan_no'):
            elements.append(Paragraph(f"PAN: {society_info['pan_no']}", center_details_style))

        elements.append(Spacer(1, 0.2*inch))

        # Receipt title and details in a 2-column layout
        date_str = voucher_data.get('date', 'N/A')
        rv_no = voucher_data.get('voucher_number', 'N/A')
        ref_val = voucher_data.get('reference', 'N/A')

        # Left side: Receipt title and labels
        labels_info = [
            [Paragraph("<b>Date:</b>", label_style), Paragraph(date_str, val_style)],
            [Paragraph("<b>Receipt No:</b>", label_style), Paragraph(rv_no, val_style)],
            [Paragraph("<b>Ref:</b>", label_style), Paragraph(ref_val, val_style)],
        ]
        labels_table = Table(labels_info, colWidths=[1.1*inch, 1.4*inch])
        labels_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ]))

        # Header row with Receipt title and labels
        header_row = Table([
            [Paragraph("Receipt", title_style), ""]
        ], colWidths=[3*inch, 3.5*inch])
        header_row.setStyle(TableStyle([
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('LINEBELOW', (0,0), (-1,0), 1, colors.grey),
            ('BOTTOMPADDING', (0,0), (-1,0), 10),
        ]))

        elements.append(header_row)
        elements.append(Spacer(1, 0.1*inch))
        elements.append(labels_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Summary Details (More formal with labels)
        member_name = voucher_data.get('member_name', 'Member')
        flat_number = voucher_data.get('flat_number', 'N/A')
        amount = voucher_data.get('total_debit', 0)
        amount_words = voucher_data.get('amount_in_words', '')
        description = voucher_data.get('description', 'Monthly Maintenance Dues')
        
        # Using Rs. symbol for PDF compatibility
        rupee_symbol = "Rs."

        details_table_data = [
            [Paragraph("<b>Received From:</b>", label_style), Paragraph(member_name, val_style)],
            [Paragraph("<b>Flat No:</b>", label_style), Paragraph(flat_number if flat_number != 'N/A' else '-', val_style)],
            [Paragraph("<b>A Sum of:</b>", label_style), Paragraph(f"<b>{rupee_symbol} {amount:,.2f}</b> ({amount_words})", val_style)],
            [Paragraph("<b>Towards:</b>", label_style), Paragraph(description, val_style)],
        ]
        
        details_table = Table(details_table_data, colWidths=[1.5*inch, 4*inch])
        details_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ]))
        elements.append(details_table)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Table of Details
        grid_data = [['Description', f'Amount ({rupee_symbol})']]
        # For a receipt, we usually show the credit entries (what was paid for)
        for entry in voucher_data.get('entries', []):
            if entry.get('credit', 0) > 0:
                grid_data.append([
                    entry.get('account_name', ''),
                    f"{float(entry.get('credit', 0)):,.2f}"
                ])

        # Total row
        grid_data.append(['Total Received', f"{rupee_symbol} {amount:,.2f}"])
        
        grid_table = Table(grid_data, colWidths=[4*inch, 1.5*inch])
        grid_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f8e9')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2e7d32')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('GRID', (0, 0), (-1, -2), 0.5, colors.grey),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(grid_table)
        elements.append(Spacer(1, 0.4 * inch))
        
        # Footer: Bank Details & Signature
        bank_info = [
            [Paragraph("<b>Payment Ref:</b>", label_style), Paragraph(voucher_data.get('reference', 'N/A'), val_style)],
        ]
        if society_info.get('bank_name'):
             bank_info.append([Paragraph("<b>Bank:</b>", label_style), Paragraph(f"{society_info['bank_name']} - {society_info['bank_account_number']}", val_style)])
             bank_info.append([Paragraph("<b>IFSC:</b>", label_style), Paragraph(society_info['bank_ifsc_code'], val_style)])

        bank_table = Table(bank_info, colWidths=[1.5*inch, 3*inch])
        bank_table.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ]))
        
        footer_data = [
            [bank_table, Paragraph("<b>Authorised Signatory</b><br/><br/><br/>_________________", ParagraphStyle('Sign', alignment=TA_RIGHT))]
        ]
        footer_table = Table(footer_data, colWidths=[4*inch, 2*inch])
        footer_table.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'BOTTOM'),
        ]))
        elements.append(footer_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_voucher_pdf(
        voucher_data: Dict[str, Any],
        society_info: Dict[str, Any]
    ) -> BytesIO:
        """
        Create a professional accounting voucher PDF
        Works for Receipt, Payment, and Journal vouchers
        """
        v_type = voucher_data.get('voucher_type', 'VOUCHER').upper()
        if v_type == 'RECEIPT':
            return PDFExporter.create_formal_receipt_pdf(voucher_data, society_info)
            
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=40,
            leftMargin=40,
            topMargin=40,
            bottomMargin=40
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'VoucherTitle',
            parent=styles['Heading1'],
            fontSize=22,
            textColor=colors.HexColor('#1f4788'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        info_style = ParagraphStyle(
            'VoucherInfo',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        
        v_type = voucher_data.get('voucher_type', 'VOUCHER').upper()
        
        # Determine Voucher Color based on type
        voucher_color = colors.black
        if v_type == 'RECEIPT':
            voucher_color = colors.HexColor('#2e7d32') # Green
        elif v_type == 'PAYMENT':
            voucher_color = colors.HexColor('#c62828') # Red
        elif v_type == 'JOURNAL':
            voucher_color = colors.HexColor('#1565c0') # Blue

        type_style = ParagraphStyle(
            'VoucherType',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=voucher_color,
            alignment=TA_CENTER,
            spaceBefore=20,
            spaceAfter=20,
            fontName='Helvetica-Bold'
        )

        # Header with Logo
        logo_url = society_info.get('logo_url')
        if logo_url:
            try:
                # Add logo if possible (simplified for now, usually requires Image flowable)
                elements.append(Paragraph(f"<img src='{logo_url}' width='60' height='60' />", title_style))
                elements.append(Spacer(1, 0.2 * inch))
            except:
                pass

        elements.append(Paragraph(society_info.get('name', 'Society Name'), title_style))
        if society_info.get('address'):
            elements.append(Paragraph(society_info['address'], info_style))
        elements.append(Spacer(1, 0.1 * inch))
        
        elements.append(Paragraph(f"{v_type} VOUCHER", type_style))
        
        # Main Info Table
        main_info = [
            [f"Voucher No: {voucher_data.get('voucher_number', 'N/A')}", f"Date: {voucher_data.get('date', 'N/A')}"],
            ["Reference:", voucher_data.get('reference', 'N/A')],
        ]
        
        info_table = Table(main_info, colWidths=[3*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Description/Narration - aligned to left properly
        narration_style = ParagraphStyle(
            'NarrationStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT,
            leftIndent=0,
            spaceBefore=0,
            spaceAfter=0
        )
        desc_para = Paragraph(f"<b>Narration:</b> {voucher_data.get('description', 'N/A')}", narration_style)
        elements.append(desc_para)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Grid Table
        rupee_symbol = "Rs."
        grid_data = [['Account Code', 'Account Name', f'Debit ({rupee_symbol})', f'Credit ({rupee_symbol})']]
        for entry in voucher_data.get('entries', []):
            grid_data.append([
                entry.get('account_code', ''),
                entry.get('account_name', ''),
                f"{float(entry.get('debit', 0)):,.2f}" if entry.get('debit', 0) > 0 else '',
                f"{float(entry.get('credit', 0)):,.2f}" if entry.get('credit', 0) > 0 else ''
            ])
            
        # Total row
        grid_data.append([
            '', 'TOTAL',
            f"{float(voucher_data.get('total_debit', 0)):,.2f}",
            f"{float(voucher_data.get('total_credit', 0)):,.2f}"
        ])
        
        grid_table = Table(grid_data, colWidths=[1*inch, 2.5*inch, 1.25*inch, 1.25*inch])
        grid_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f0f0f0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ])
        grid_table.setStyle(grid_style)
        elements.append(grid_table)
        elements.append(Spacer(1, 0.5 * inch))
        
        # Signatures
        sig_data = [['Prepared By', 'Checked By', 'Approved By']]
        sig_table = Table(sig_data, colWidths=[2*inch, 2*inch, 2*inch])
        sig_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica', 10),
            ('LINEABOVE', (0, 0), (-1, 0), 1, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
        ]))
        elements.append(sig_table)
        
        doc.build(elements)
        buffer.seek(0)
        return buffer
